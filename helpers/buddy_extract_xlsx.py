#!/usr/bin/env python3
"""Extract structured data from XLSX files in buddy-db.

Downloads XLSX from OneDrive, parses with openpyxl, stores in raw_structured.
Also scans for birthday patterns in cell values.

No LLM, no MarkItDown — pure Python parsing.
"""
import json
import os
import re
import sys
import tempfile

import httpx
import openpyxl
import psycopg2
import psycopg2.extras

BUDDY_DB = {"host": "buddy-db", "port": 5432, "dbname": "buddy", "user": "buddy", "password": "buddy"}
DRIVE_ID = "8DF5EDA212792823"

BIRTHDAY_PATTERNS = re.compile(
    r'\b(geb\.?|geburtstag|birthday|geboren|geb[\s\-]?datum|birth[\s\-]?date|geburtsdatum)\b',
    re.IGNORECASE
)

DATE_PATTERN = re.compile(
    r'\b(\d{1,2}[./\-]\d{1,2}[./\-]\d{2,4})\b'
)


def get_download_url(source_id):
    """Get OneDrive pre-signed download URL via M365 MCP tool."""
    from brix.sdk import mcp
    result = mcp.call("m365", "download-onedrive-file-content", {
        "driveId": DRIVE_ID,
        "driveItemId": source_id,
    })
    if isinstance(result, dict) and result.get("success"):
        data = result.get("data", {})
        if isinstance(data, dict):
            return data.get("@microsoft.graph.downloadUrl")
        if isinstance(data, str):
            try:
                import ast
                d = ast.literal_eval(data)
                return d.get("@microsoft.graph.downloadUrl")
            except Exception:
                pass
    return None


def download_file(client, url):
    """Download file bytes from pre-signed URL."""
    resp = client.get(url, follow_redirects=True, timeout=60.0)
    resp.raise_for_status()
    return resp.content


def cell_to_value(cell):
    """Convert openpyxl cell to JSON-serializable value."""
    if cell.value is None:
        return None
    if cell.data_type == 'f':
        return f"={cell.value}"
    if isinstance(cell.value, (int, float, bool)):
        return cell.value
    if hasattr(cell.value, 'isoformat'):
        return cell.value.isoformat()
    return str(cell.value)


def extract_sheets(wb):
    """Extract all sheets from workbook as structured data."""
    sheets = []
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_data = []
        for row in ws.iter_rows():
            row_values = [cell_to_value(c) for c in row]
            if any(v is not None for v in row_values):
                rows_data.append(row_values)

        sheet_info = {
            "name": sheet_name,
            "rows": len(rows_data),
            "cols": ws.max_column or 0,
            "data": rows_data,
        }
        total_rows += len(rows_data)
        sheets.append(sheet_info)

    return {
        "sheets": sheets,
        "total_sheets": len(sheets),
        "total_rows": total_rows,
    }


def scan_birthday_hints(structured_data):
    """Scan all cells for birthday-related patterns."""
    hints = []

    for sheet in structured_data.get("sheets", []):
        sheet_name = sheet["name"]
        data = sheet.get("data", [])

        for row_idx, row in enumerate(data):
            for col_idx, val in enumerate(row):
                if val is None or not isinstance(val, str):
                    continue

                if BIRTHDAY_PATTERNS.search(val):
                    hint = {
                        "keyword": val.strip()[:100],
                        "sheet": sheet_name,
                        "row": row_idx + 1,
                        "col": col_idx + 1,
                    }

                    # Search for name (first non-empty text cell in same row that is not keyword/date)
                    name_candidates = []
                    for c_idx, c_val in enumerate(row):
                        if c_val and isinstance(c_val, str) and c_idx != col_idx:
                            if not BIRTHDAY_PATTERNS.search(c_val) and not DATE_PATTERN.search(c_val):
                                name_candidates.append(c_val.strip())
                    if name_candidates:
                        hint["name"] = name_candidates[0][:100]

                    # Search for dates in same row
                    for c_idx, c_val in enumerate(row):
                        if c_val and isinstance(c_val, str):
                            date_match = DATE_PATTERN.search(c_val)
                            if date_match:
                                hint["date"] = date_match.group(1)
                                break
                            if 'T' in c_val and len(c_val) >= 10:
                                hint["date"] = c_val[:10]
                                break

                    hints.append(hint)

    return hints


def fetch_xlsx_docs(conn, folder_path_pattern=None, batch_ref=None, source_pattern=None, limit=100):
    """Fetch pending XLSX documents from buddy-db."""
    conditions = ["extension = 'xlsx'", "processing_status = 'pending'"]
    params = []

    if folder_path_pattern:
        conditions.append("folder_path LIKE %s")
        params.append(folder_path_pattern)

    if batch_ref:
        conditions.append("batch_ref = %s")
        params.append(batch_ref)

    if source_pattern:
        conditions.append("source LIKE %s")
        params.append(source_pattern)

    query = f"""
        SELECT id, file_name, source_id, folder_path, file_size
        FROM documents
        WHERE {' AND '.join(conditions)}
        ORDER BY id
        LIMIT %s
    """
    params.append(int(limit))

    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(query, params)
        return cur.fetchall()


def update_document(conn, doc_id, structured_data, birthday_hints):
    """Update document with extracted data."""
    with conn.cursor() as cur:
        # Store structured data + mark as done + append specialist
        cur.execute("""
            UPDATE documents
            SET raw_structured = %s,
                processing_status = 'done',
                processed_at = NOW(),
                extraction_specialists = array_append(
                    COALESCE(extraction_specialists, ARRAY[]::text[]),
                    'xlsx_direct'
                )
            WHERE id = %s
              AND NOT ('xlsx_direct' = ANY(COALESCE(extraction_specialists, ARRAY[]::text[])))
        """, (json.dumps(structured_data), doc_id))

        # If xlsx_direct already in array, just update the other fields
        if cur.rowcount == 0:
            cur.execute("""
                UPDATE documents
                SET raw_structured = %s,
                    processing_status = 'done',
                    processed_at = NOW()
                WHERE id = %s
            """, (json.dumps(structured_data), doc_id))

        # Store extraction in document_extractions
        extraction_data = {
            "total_sheets": structured_data["total_sheets"],
            "total_rows": structured_data["total_rows"],
            "sheet_names": [s["name"] for s in structured_data["sheets"]],
            "birthday_hints": birthday_hints,
        }

        cur.execute("""
            INSERT INTO document_extractions (document_id, specialist_name, extraction_data, confidence, extracted_at)
            VALUES (%s, 'xlsx_direct', %s, 1.0, NOW())
            ON CONFLICT (document_id, specialist_name) DO UPDATE
            SET extraction_data = EXCLUDED.extraction_data,
                confidence = EXCLUDED.confidence,
                extracted_at = EXCLUDED.extracted_at
        """, (doc_id, json.dumps(extraction_data)))

    conn.commit()


def mark_error(conn, doc_id, error_msg):
    """Mark document as error."""
    with conn.cursor() as cur:
        cur.execute("""
            UPDATE documents
            SET processing_status = 'error',
                processing_error = %s,
                processed_at = NOW()
            WHERE id = %s
        """, (error_msg[:500], doc_id))
    conn.commit()


def main():
    if len(sys.argv) > 1:
        arg = sys.argv[1]
        if arg.startswith('/') and os.path.isfile(arg):
            with open(arg) as f:
                params = json.load(f)
        else:
            params = json.loads(arg)
    elif not sys.stdin.isatty():
        raw = sys.stdin.read().strip()
        params = json.loads(raw) if raw else {}
    else:
        params = {}

    folder_path_pattern = params.get("folder_path_pattern")
    batch_ref = params.get("batch_ref")
    source_pattern = params.get("source_pattern")
    limit = params.get("limit", "100")

    conn = psycopg2.connect(**BUDDY_DB)
    client = httpx.Client(timeout=120.0)

    docs = fetch_xlsx_docs(conn, folder_path_pattern, batch_ref, source_pattern, limit)
    total = len(docs)
    processed = 0
    errors = 0
    total_birthday_hints = 0
    details = []

    sys.stderr.write(f"[buddy_extract_xlsx] Found {total} pending XLSX documents\n")

    for i, doc in enumerate(docs):
        doc_id = doc["id"]
        file_name = doc["file_name"]
        source_id = doc["source_id"]
        detail = {"id": doc_id, "file_name": file_name, "status": "error"}

        try:
            # Step 1: Get download URL
            url = get_download_url(source_id)
            if not url:
                raise Exception("No download URL from M365")

            # Step 2: Download file
            file_bytes = download_file(client, url)

            # Step 3: Parse with openpyxl
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=True) as tmp:
                tmp.write(file_bytes)
                tmp.flush()
                wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
                structured_data = extract_sheets(wb)
                wb.close()

            # Step 4: Scan for birthday hints
            birthday_hints = scan_birthday_hints(structured_data)
            if birthday_hints:
                structured_data["birthday_hints"] = birthday_hints
                total_birthday_hints += len(birthday_hints)

            # Step 5: Update DB
            update_document(conn, doc_id, structured_data, birthday_hints)
            processed += 1
            detail["status"] = "ok"
            detail["sheets"] = structured_data["total_sheets"]
            detail["rows"] = structured_data["total_rows"]
            if birthday_hints:
                detail["birthday_hints"] = len(birthday_hints)

            if (i + 1) % 10 == 0:
                sys.stderr.write(f"[buddy_extract_xlsx] Progress: {i+1}/{total} (ok={processed}, err={errors})\n")
                sys.stderr.write(f"BRIX_PROGRESS:{int((i+1)/total*100)}\n")
                sys.stderr.flush()

        except Exception as e:
            errors += 1
            err_msg = str(e)[:200]
            detail["reason"] = err_msg
            mark_error(conn, doc_id, err_msg)
            sys.stderr.write(f"[buddy_extract_xlsx] Error {file_name}: {e}\n")

        details.append(detail)

    conn.close()
    client.close()

    sys.stderr.write(f"[buddy_extract_xlsx] Done: {processed}/{total} ok, {errors} errors, {total_birthday_hints} birthday hints\n")

    print(json.dumps({
        "total": total,
        "processed": processed,
        "errors": errors,
        "birthday_hints_found": total_birthday_hints,
        "details": details[:100],
    }))


if __name__ == "__main__":
    main()
